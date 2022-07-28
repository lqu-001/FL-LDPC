#update from merge_dnbit8_dnberbyiter12.py, this can calculate avg iterations

#param_num=5.82E5, total_bits=5.82E5*8=4.65E6, frame_number =5.82E5*8/1008=4620
#np.random.choice(a,N,replace=False)#only apply to 1-D sequence; time unchange
#random.sample(a,N) #can apply to different dimension, default not replaced; time consume more than np when N is large than 10000
#torch.rand:unifrom; torch.randn:std
from tkinter import _flatten
from openpyxl import Workbook #xls
import matplotlib
matplotlib.use('TkAgg')
import matplotlib.pyplot as plt
import copy
import numpy as np
np.set_printoptions(threshold=np.inf)
from torchvision import datasets, transforms
import torch
import math
import random #uniform distribution, rather than gaussian

from utils.sampling import mnist_iid, mnist_noniid, fashionmnist_iid, fashionmnist_noniid, cifar_iid
from utils.options import args_parser
from models.Update import LocalUpdate
from models.Nets import lenet5,vanillacnn, cnn3, resnet18
from models.Fed import FedAvg
from models.mask import *
from models.test import test_img

import time
start=time.time()

args = args_parser()

seed=1
np.random.seed(seed)
torch.manual_seed(seed)
torch.cuda.manual_seed_all(seed)
torch.backends.cudnn.deterministic = True
torch.backends.cudnn.benchmark = False

#-------------ldpc--------------------------------------
import encode
import modulate
import decode
from utility import *
import pickle
import sys
def ldpc_decoder(s, max_it):
    MATRIX_PATH = "Matrix(2016,1008)Block56.mat"
    Hp, Hs, H = load_H(MATRIX_PATH)    
    SNR = 2.5
    sigma2 = 1.0 / pow(10.0, SNR / 10.0)
    #s = np.random.randint(0, 2, (1, K), dtype=np.int)#s=[[1,1,0,0,1]]; type() is numpy.ndarray; .shape is (1,1008); .dtype is int64
    x = encode.method_2(s, Hp, Hs)
    y = modulate.awgn(modulate.BPSK(x), SNR)
    y_llr = decode.LLR(y, sigma2).reshape(-1)
    x_hat, frameiter = decode.MS(y_llr, H, max_it)
    BERcnt = np.sum(np.abs(s - x_hat))
    if (BERcnt == 0):
        FCRcnt = 1
    else:
        FCRcnt = 0 
    return x_hat, BERcnt, FCRcnt, frameiter
#-----------------ldpc------------------------------------------

if __name__ == '__main__':    
    args.device = torch.device('cuda:{}'.format(args.gpu) if torch.cuda.is_available() and args.gpu != -1 else 'cpu')        
#dataset---------------------------------------------------
    if args.dataset == 'fashionmnist': 
        transform = transforms.Compose([transforms.ToTensor(), transforms.Normalize((0.5,), (0.5,))]) 
        dataset_train = datasets.FashionMNIST('./data/FashionMNIST', download=True, train=True, transform=transform)
        dataset_test = datasets.FashionMNIST('./data/FashionMNIST', download=True, train=False, transform=transform)
        if args.iid:
            dict_users = fashionmnist_iid(dataset_train, args.num_users)
        else:
            dict_users = fashionmnist_noniid(dataset_train, args.num_users)
            
    elif args.dataset == 'cifar10':        
        transform1 = transforms.Compose([transforms.RandomHorizontalFlip(),transforms.RandomGrayscale(),transforms.ToTensor(),transforms.Normalize((0.4914, 0.4822, 0.4465), (0.2023, 0.1994, 0.2010))])
        transform2 = transforms.Compose([transforms.ToTensor(), transforms.Normalize((0.4914, 0.4822, 0.4465), (0.2023, 0.1994, 0.2010))]) #(0.4914, 0.4822, 0.4465), (0.2023, 0.1994, 0.2010)
        dataset_train = datasets.CIFAR10('./data/cifar', train=True, download=True, transform=transform1)#(0.5, 0.5, 0.5), (0.5, 0.5, 0.5)
        dataset_test = datasets.CIFAR10('./data/cifar', train=False, download=True, transform=transform2)
        if args.iid:
            dict_users = cifar_iid(dataset_train, args.num_users)
        else:
            #exit('Error: only consider IID setting in CIFAR10')
            dict_users = cifar_noniid(dataset_train, args.num_users)
    else:
        exit('Error: unrecognized dataset')
    img_size = dataset_train[0][0].shape  
 
#model-------------------------------------------------------    
    if args.model == 'vanillacnn' and args.dataset == 'fashionmnist':
        net_server = vanillacnn(args=args).to(args.device)
    elif args.model == 'resnet18' and args.dataset == 'cifar10':
        net_server = resnet18().to(args.device)
    elif args.model == 'lenet5' and args.dataset == 'fashionmnist':
        net_server = lenet5(args=args).to(args.device)
    else:
        exit('Error: unrecognized model')
    #print(net_server) 
    net_server.train()    

    for key in net_server.state_dict().keys():
        print(key, net_server.state_dict()[key].size())
#server to clients-------------------------------------------    
    glob_acc, glob_train_loss = [],[]
    wb = Workbook() #xls
    ws = wb.active  #xls
    bitsvolume=0
    avg_losss=[]
#------------rounds------------------------------------------  
    for rounds in range(args.rounds):
        dnbit = args.dnbit
        serverbroad = copy.deepcopy(net_server.state_dict())
        #print('---------------------server broadcast-----------------------')
        #print(serverbroad['fc1.weight'][:5,:10])  
        wglob = copy.deepcopy(net_server.state_dict())
        tmp = copy.deepcopy(net_server.state_dict())
        wldpc = copy.deepcopy(net_server.state_dict())##
        net_client_in = copy.deepcopy(net_server.to(args.device))##        
        dw_locals=[]; loss_locals = []   
#-------------------------------------------------------------
        '''print('type(conv1.weight):',type(wldpc['conv1.weight'])) #torch.Tenssor
        print('conv1.weight.dtype:',wldpc['conv1.weight'].dtype) #torch.float32
        print('wldpc[conv1.weight].size:',wldpc['conv1.weight'].size()) #torch.Size([32, 1, 5, 5])
        print('wldpc[conv1.weight].shape:',wldpc['conv1.weight'].shape) #torch.Size([32, 1, 5, 5])
        print('len(wldpc[conv1.weight]):',len(wldpc['conv1.weight'])) #32
        print('wldpc[conv1.weight].numel():',wldpc['conv1.weight'].numel()) #800 '''
        param_num = sum(wldpc[i].numel() for i in wldpc.keys()) #582026
        #print('param_num',param_num) #582026
        
#----------------get fp32 1d-sequence: seqfp32---------------------------------        
        seqfp32 = wldpc['conv1.weight'].flatten() #flatten horizontal
        for key in wldpc.keys():
            if (key != 'conv1.weight'):
                seqfp32 = torch.cat((seqfp32, wldpc[key].flatten()), -1) #no matter -1 or 0, torch.Size([11183582]); cat(-1) horizontal          
        #print('type(seqfp32):',type(seqfp32)) #torch.Tensor
        #print('seqfp32.dtype:',seqfp32.dtype) #torch.float32
        #print('seqfp32.shape:',seqfp32.shape) #torch.Size([582026])
        
        seqbin, factor, mini = quan_bit(seqfp32, dnbit)
        #get binary sequence seqbin=['1001','0010'...]
        #print('type(seqbin1)',type(seqbin)) #list
        #print('seqbin1.dtype',seqbin.dtype)  #list has no attribute dtype
        
        for j in range (len(seqbin)):
            seqbin[j]=list(seqbin[j]) #change to list seqbin=[['1','0','0','1'],['0','0','1','0']...]
        #print('type(seqbin2)',type(seqbin)) #list
        #print('seqbin2.dtype',seqbin.dtype) #list has no attribute dtype
        
#####---------------here is ldpc input---------------------------------------#######

        list1d = list(_flatten(seqbin)) #change into 1 dimension list
        #print('list1d',list1d, type(list1d)) #['1','1','0',...,'0'], list, list has no dtype
        for ri in range(len(list1d)):
            if (list1d[ri]=='1'):
                list1d[ri]=1
            else:
                list1d[ri]=0
        #print('list1d',list1d, type(list1d)) #[1,1,0,...,0], list, list has no dtype
        array1dc=np.array(list1d)
        #print('array1dc',array1dc, type(array1dc), array1dc.shape, array1dc.dtype) #[1,1,0,...,0], np.ndarray, (4656208,), int64
        array1dr = array1dc.reshape(1,-1)
        #print('array1dr',array1dr, type(array1dr), array1dr.shape, array1dr.dtype) #[[1,1,0,...,0]],np.ndarray, (1,4656208), int64
        
        tailbit=1008-(param_num*dnbit)%1008
        tail = np.zeros((1,tailbit), dtype = np.int)
        ldpcin=np.concatenate((array1dr,tail),axis=1)
        #print(ldpcin, 'ldpcin', type(ldpcin), ldpcin.shape, ldpcin.dtype) #[[1,1,0,...,0]],np.ndarray, (1,4656960), int64        
        ldpcin=ldpcin.reshape(-1,1008)
        #print(ldpcin, 'ldpcin', type(ldpcin), ldpcin.shape, ldpcin.dtype) ndarray, (4620,1008), int64
                        
        MaxIter=int(np.ceil(rounds*12/50+12))
        ww=ldpcin[0,:]
        ww=ww.reshape(1,-1)
        ldpcout,BERcnt,FCRcnt,frameiter=ldpc_decoder(ww,MaxIter) #get the 1st base, later sequence can be added on
        frame_number=int(np.ceil(param_num*dnbit/1008))
        for i in range(1,frame_number):
            ww0=ldpcin[i,:].reshape(1,-1)
            ldpcoutmp,BERcntmp,FCRcntmp,frameitertmp=ldpc_decoder(ww0,MaxIter)
            ldpcout=np.concatenate((ldpcout,ldpcoutmp),axis=1)
            BERcnt+=BERcntmp
            FCRcnt+=FCRcntmp
            frameiter+=frameitertmp #accumulate all frames' iters
            print('frameiter ',frameitertmp)
        BER=BERcnt/(1008*np.ceil(param_num*dnbit/1008))
        FER=1-FCRcnt/np.ceil(param_num*dnbit/1008)
        AvgIter=frameiter/frame_number

        ldpcout=ldpcout[0,0:param_num*dnbit] #cut tail [[1,1,0,...,0]]
        ldpcout1=ldpcout.reshape(-1,dnbit)
        #print(ldpcout1,'ldpcout1',type(ldpcout1), ldpcout1.shape, ldpcout1.dtype) #[[1,1][0,1]],np.ndarray, (582026,8), int64
        seqbin=ldpcout1.tolist() #print(ldpcout1,'ldpcout1',type(ldpcout1))#[[1,1],[0,1]]
        for i in range(param_num):
            for j in range(dnbit):
                if (seqbin[i][j]==1):
                    seqbin[i][j]='1'
                else:
                    seqbin[i][j]='0'
        #print(seqbin,'seqbin',type(seqbin))#[['1','1'],['0','1']]

######--------------here is ldpc output--------------------------------------#######
        
        for k in range (len(seqbin)):
            seqbin[k]=''.join(seqbin[k])    # seqbin=['0001','0010'...]
        #print('type(seqbin4)',type(seqbin)) #list
        #print('seqbin4.dtype',seqbin.dtype) #list has no attribute dtype
        seqfp32_ber = dequan_bit(seqbin,factor, mini, dnbit) #seqfp32_ber=[0.32,-0.02,...]
        seqfp32_ber = torch.Tensor(seqfp32_ber).to(args.device)
        #print('type(seqfp32_ber)',type(seqfp32_ber)) #torch.Tensor
        #print('seqfp32_ber.dtype',seqfp32_ber.dtype) #torch.float32
        
#-----------find out the error elements-----------------------------  
        '''#print('****************************before correction******************************************')      
        broadcast=[]; receive=[]; difference=[]; diff_location=[]
        for i in range (len(seqfp32_ber)):
            if (seqfp32_ber[i]!=seqfp32[i]): #not work, because quant dequant itself will cause quantization error, every element is different
                broadcast.append(seqfp32[i])
                receive.append(seqfp32_ber[i])
                difference.append(seqfp32_ber[i]-seqfp32[i])
                diff_location.append(i)
        #print('-------------------broadcast--------------------', '\n', broadcast)
        #print('-------------------receive----------------------', '\n', receive)
        #print('-----------------difference---------------------', '\n', difference)
        #print('---------------diff_location--------------------', '\n', diff_location)                
        #-----demo-----
        print('-------------------affected weight----------------------------')
        print('Model size is:', param_num, '; Model bits is:', param_num*dnbit, '; BER is:', args.ber)
        print('Affected weight number is: ', len(diff_location))        
        #print('-----difference sort-----', '\n', sorted(difference))
        difference_sort, idx = torch.tensor(difference).sort(dim=-1)
        location_sort=[]
        for i in idx:
            location_sort.append(diff_location[i])
        #print(difference_sort)
        #print(location_sort)        
        pair=[]
        for i in range(len(location_sort)):
            pair.append((difference_sort[i],location_sort[i]))
        print('---------------sorted difference and location--------------------', '\n', pair)

        #print('-------------------location map to layer--------------------')
        #print('Conv1 [0:831]; Conv2 [832:52095]; FC1 [52096:576895]; FC2 [576896:582025]')'''
        
#-----------rehape and load the noised param to NN------------------------------        
        coodi=0
        for p in wldpc.keys():
            coodi += wldpc[p].numel()
            layer_ber = (seqfp32_ber[coodi-wldpc[p].numel():coodi]).reshape(wldpc[p].shape)
            net_client_in.state_dict()[p].data.copy_(layer_ber)
        net_client_in.to(args.device)
        net_client_in.eval()
        
        #print('---------------------client receive----------------------')
        #print(net_client_in.state_dict()['fc1.weight'][:5,:10])
        #print('---------------------difference----------------------')
        #print(net_client_in.state_dict()['fc1.weight'][:5,:10]-serverbroad['fc1.weight'][:5,:10])                
                
#-----------local train-----------------------------------
        client_join = max(int(args.frac * args.num_users), 1)
        idxs_users = np.random.choice(args.num_users, client_join, replace=False) #[0,3,5,9,2,...]; np.random: equal problity; False means not repeated        
        for idx in idxs_users:         
            local = LocalUpdate(args=args, dataset=dataset_train, idxs=dict_users[idx])
            w, mode, loss = local.train(net=copy.deepcopy(net_client_in).to(args.device))
            for k in w.keys():                
                dw=copy.deepcopy(w[k])-net_client_in.state_dict()[k]                   
                tmp[k].data.copy_(dw)
            dw_locals.append(copy.deepcopy(tmp))
            loss_locals.append(copy.deepcopy(loss))
        dw_glob=FedAvg(dw_locals)
#-----------server agg and train test---------------------        
        for k in wglob.keys():
            wglob[k]+=dw_glob[k]
        net_server.load_state_dict(wglob)
        
        net_server.to(args.device)
        net_server.eval()        
        test_acc, test_loss = test_img(net_server, dataset_test, args)
        train_acc, train_loss = test_img(net_server, dataset_train, args) 
        avg_loss = sum(loss_locals) / len(loss_locals)
        avg_losss.append(avg_loss)
        
        print(test_acc.numpy())
        print('Avg_loss:',avg_loss)
        print('Train_loss {:.3f}, Test_loss {:.3f}'.format(train_loss, test_loss))
        print('BER: ',BER, ';     AvgIter: ', AvgIter)
        
        bitsvolume=bitsvolume+ client_join*32*param_num/(10**9) #Gb
        
        ws.cell(rounds+2,1).value = rounds  
        ws.cell(rounds+2,2).value = str(test_acc.numpy()) 
        ws.cell(rounds+2,3).value = str(test_loss)
        ws.cell(rounds+2,4).value = str(avg_loss)
        ws.cell(rounds+2,5).value = AvgIter #for one frame
        ws.cell(rounds+2,6).value = BER
        ws.cell(rounds+2,7).value = FER
        ws.cell(rounds+2,8).value = frameiter #one round, all iters for all frames
                                    
    wb.save("./result/merge_dnbit8_dnberMaxIterLinear.xlsx")
    
end=time.time()
print('running time is ',end-start)
